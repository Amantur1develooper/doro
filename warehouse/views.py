import os
import re
from datetime import date, timedelta, datetime

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.shortcuts import render, redirect, get_object_or_404

from .models import Warehouse, Product, ProductCategory, Batch, StockMovement

DATA_FOLDER = os.path.join(settings.BASE_DIR, 'дора лайн движение')


@login_required
def warehouses_list(request):
    warehouses = Warehouse.objects.all()
    return render(request, 'warehouse/warehouses_list.html', {'warehouses': warehouses})


@login_required
def warehouse_detail(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    batches = Batch.objects.filter(warehouse=warehouse, quantity__gt=0).select_related('product', 'product__category')
    movements = StockMovement.objects.filter(
        batch__warehouse=warehouse
    ).select_related('batch__product', 'employee').order_by('-date')[:20]
    expiring = batches.filter(expiry_date__lte=date.today() + timedelta(days=90))
    expired = batches.filter(expiry_date__lt=date.today())
    return render(request, 'warehouse/warehouse_detail.html', {
        'warehouse': warehouse, 'batches': batches,
        'movements': movements, 'expiring': expiring, 'expired': expired
    })


@login_required
def products_list(request):
    products = Product.objects.select_related('category')
    search = request.GET.get('q', '')
    category_id = request.GET.get('category')
    if search:
        products = products.filter(name__icontains=search)
    if category_id:
        products = products.filter(category_id=category_id)
    categories = ProductCategory.objects.all()

    # Annotate with total stock
    for p in products:
        p.total_stock = Batch.objects.filter(
            product=p, quantity__gt=0, expiry_date__gte=date.today()
        ).aggregate(total=Sum('quantity'))['total'] or 0

    return render(request, 'warehouse/products_list.html', {
        'products': products, 'categories': categories, 'search': search
    })


@login_required
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    batches = Batch.objects.filter(product=product).select_related('warehouse')
    movements = StockMovement.objects.filter(
        batch__product=product
    ).select_related('batch__warehouse', 'employee').order_by('-date')[:30]
    return render(request, 'warehouse/product_detail.html', {
        'product': product, 'batches': batches, 'movements': movements
    })


@login_required
def product_create(request):
    categories = ProductCategory.objects.all()
    if request.method == 'POST':
        sku = request.POST.get('sku', '').strip()
        name = request.POST.get('name', '').strip()

        if not name:
            messages.error(request, 'Название товара обязательно')
            return render(request, 'warehouse/product_form.html', {'categories': categories, 'post': request.POST})

        if not sku:
            messages.error(request, 'Артикул (SKU) обязателен')
            return render(request, 'warehouse/product_form.html', {'categories': categories, 'post': request.POST})

        if Product.objects.filter(sku=sku).exists():
            messages.error(request, f'Товар с артикулом «{sku}» уже существует. Укажите другой артикул.')
            return render(request, 'warehouse/product_form.html', {'categories': categories, 'post': request.POST})

        p = Product(
            name=name,
            international_name=request.POST.get('international_name', ''),
            form=request.POST.get('form', ''),
            dosage=request.POST.get('dosage', ''),
            manufacturer=request.POST.get('manufacturer', ''),
            sku=sku,
            unit=request.POST.get('unit', 'шт'),
            purchase_price=request.POST.get('purchase_price', 0) or 0,
            cost_price=request.POST.get('cost_price', 0) or 0,
            sale_price=request.POST.get('sale_price', 0) or 0,
            notes=request.POST.get('notes', ''),
        )
        cat_id = request.POST.get('category')
        if cat_id:
            p.category_id = cat_id
        p.save()
        messages.success(request, f'Товар «{name}» добавлен')
        return redirect('products_list')
    return render(request, 'warehouse/product_form.html', {'categories': categories})


@login_required
def warehouse_create(request):
    from accounts.models import User
    managers = User.objects.filter(role__in=['warehouse', 'sales_manager', 'superadmin', 'director'])
    if request.method == 'POST':
        w = Warehouse(
            name=request.POST.get('name', '').strip(),
            address=request.POST.get('address', '').strip(),
        )
        manager_id = request.POST.get('manager')
        if manager_id:
            w.manager_id = manager_id
        w.save()
        messages.success(request, f'Склад «{w.name}» добавлен')
        return redirect('warehouses_list')
    return render(request, 'warehouse/warehouse_form.html', {'managers': managers})


@login_required
def stock_in(request):
    warehouses = Warehouse.objects.all()
    products = Product.objects.all()
    if request.method == 'POST':
        batch = Batch(
            product_id=request.POST.get('product'),
            batch_number=request.POST.get('batch_number'),
            expiry_date=request.POST.get('expiry_date'),
            warehouse_id=request.POST.get('warehouse'),
            quantity=int(request.POST.get('quantity', 0)),
            purchase_price=request.POST.get('purchase_price', 0),
        )
        batch.save()
        StockMovement.objects.create(
            movement_type='in',
            batch=batch,
            warehouse_to=batch.warehouse,
            quantity=batch.quantity,
            price=batch.purchase_price,
            employee=request.user,
            notes=request.POST.get('notes', ''),
        )
        messages.success(request, f'Товар принят: {batch.product.name} × {batch.quantity}')
        return redirect('warehouses_list')
    return render(request, 'warehouse/stock_in.html', {
        'warehouses': warehouses, 'products': products
    })


@login_required
def stock_writeoff(request):
    warehouses = Warehouse.objects.all()
    products = Product.objects.all()
    if request.method == 'POST':
        batch = get_object_or_404(Batch, pk=request.POST.get('batch'))
        qty = int(request.POST.get('quantity', 0))
        if qty > batch.quantity:
            messages.error(request, 'Недостаточно товара')
        else:
            batch.quantity -= qty
            batch.save()
            StockMovement.objects.create(
                movement_type='writeoff',
                batch=batch,
                warehouse_from=batch.warehouse,
                quantity=qty,
                employee=request.user,
                notes=request.POST.get('notes', ''),
            )
            messages.success(request, 'Товар списан')
        return redirect('warehouses_list')
    batches = Batch.objects.filter(quantity__gt=0).select_related('product', 'warehouse')
    return render(request, 'warehouse/stock_writeoff.html', {'batches': batches})


@login_required
def expiring_report(request):
    today = date.today()
    soon = today + timedelta(days=90)
    expiring = Batch.objects.filter(
        expiry_date__lte=soon, quantity__gt=0
    ).select_related('product', 'warehouse').order_by('expiry_date')
    expired = Batch.objects.filter(
        expiry_date__lt=today, quantity__gt=0
    ).select_related('product', 'warehouse').order_by('expiry_date')
    return render(request, 'warehouse/expiring_report.html', {
        'expiring': expiring, 'expired': expired, 'today': today
    })


@login_required
def movements_list(request):
    movements = StockMovement.objects.select_related(
        'batch__product', 'batch__warehouse', 'employee'
    ).order_by('-date')[:200]
    return render(request, 'warehouse/movements_list.html', {'movements': movements})


# ------------------------------------------------------------------ #
# Excel import views (дора лайн движение)                             #
# ------------------------------------------------------------------ #

def _clean(v):
    if v is None:
        return ''
    return str(v).strip()


def _int_val(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _parse_doc_date(doc_str):
    m = re.search(r'от\s+(\d{2}\.\d{2}\.\d{4})', doc_str)
    if m:
        try:
            return datetime.strptime(m.group(1), '%d.%m.%Y').date()
        except ValueError:
            pass
    return None


def _read_summary():
    """Parse Доро Лайн движ.xlsx → list of product summary rows."""
    try:
        import openpyxl
    except ImportError:
        return [], 'openpyxl не установлен'

    path = os.path.join(DATA_FOLDER, 'Доро Лайн движ.xlsx')
    if not os.path.exists(path):
        return [], f'Файл не найден: {path}'

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    results = []
    for row in rows:
        name = _clean(row[3]) if len(row) > 3 else ''
        if not name or name in ('Номенклатура', 'Склад', 'Артикул') or name.startswith('Итого'):
            continue
        results.append({
            'product': name,
            'unit': _clean(row[8]) if len(row) > 8 else 'шт',
            'opening': _int_val(row[9]) if len(row) > 9 else 0,
            'incoming': _int_val(row[13]) if len(row) > 13 else 0,
            'outgoing': _int_val(row[15]) if len(row) > 15 else 0,
            'closing': _int_val(row[16]) if len(row) > 16 else 0,
        })
    return results, None


def _read_details():
    """Parse Доро Лайн движ детали.xlsx → list of transaction rows."""
    try:
        import openpyxl
    except ImportError:
        return [], 'openpyxl не установлен'

    path = os.path.join(DATA_FOLDER, 'Доро Лайн движ детали.xlsx')
    if not os.path.exists(path):
        return [], f'Файл не найден: {path}'

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    results = []
    current_product = None
    for row in rows:
        if len(row) < 14:
            continue
        product_name = _clean(row[12]) if row[12] else ''
        doc_name = _clean(row[0]) if row[0] else ''

        if product_name and not doc_name:
            current_product = product_name
            continue

        if doc_name and current_product and (
            doc_name.startswith('Расходный') or doc_name.startswith('Приходный')
        ):
            partner = _clean(row[10]) if len(row) > 10 else ''
            qty_in = _int_val(row[15]) if len(row) > 15 else 0
            qty_out = _int_val(row[16]) if len(row) > 16 else 0
            results.append({
                'product': current_product,
                'doc': doc_name[:80],
                'date': _parse_doc_date(doc_name),
                'partner': partner,
                'qty_before': _int_val(row[14]) if len(row) > 14 else 0,
                'qty_in': qty_in,
                'qty_out': qty_out,
                'qty_after': _int_val(row[17]) if len(row) > 17 else 0,
                'type': 'in' if qty_in else 'out',
            })
    return results, None


def _read_movements_xls():
    """Parse Движение номен-ры по фирме ИП Арапова за февраль.xls"""
    try:
        import xlrd
    except ImportError:
        return [], 'xlrd не установлен. Выполните: pip install xlrd'

    path = os.path.join(DATA_FOLDER, 'Движение номен-ры по фирме ИП Арапова за февраль.xls')
    if not os.path.exists(path):
        return [], f'Файл не найден: {path}'

    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    results = []
    for i in range(5, ws.nrows):
        row = ws.row_values(i)
        doc = _clean(row[0])
        partner = _clean(row[1])
        product = _clean(row[2])
        if not product:
            continue
        results.append({
            'doc': doc[:80],
            'partner': partner,
            'product': product,
            'incoming': _int_val(row[3]),
            'outgoing': _int_val(row[4]),
            'date': _parse_doc_date(doc),
        })
    return results, None


def _read_analysis_xls():
    """Parse Анализ номен-ры по фирме ИП Арапова за февраль.xls"""
    try:
        import xlrd
    except ImportError:
        return [], 'xlrd не установлен. Выполните: pip install xlrd'

    path = os.path.join(DATA_FOLDER, 'Анализ номен-ры по фирме ИП Арапова за февраль.xls')
    if not os.path.exists(path):
        return [], f'Файл не найден: {path}'

    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    results = []
    for i in range(6, ws.nrows):
        row = ws.row_values(i)
        product = _clean(row[1]) if len(row) > 1 else ''
        if not product or product.startswith('Итого') or product.startswith('Ед.'):
            continue
        exp_str = _clean(row[3]) if len(row) > 3 else ''
        exp_date = None
        for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
            try:
                exp_date = datetime.strptime(exp_str, fmt).date()
                break
            except ValueError:
                pass
        results.append({
            'product': product,
            'expiry_date': exp_date,
            'unit': _clean(row[4]) if len(row) > 4 else 'шт',
            'opening': _int_val(row[5]) if len(row) > 5 else 0,
            'incoming': _int_val(row[7]) if len(row) > 7 else 0,
            'outgoing': _int_val(row[9]) if len(row) > 9 else 0,
            'closing': _int_val(row[11]) if len(row) > 11 else 0,
        })
    return results, None


@login_required
def dora_report(request):
    """View: display all 4 Excel files as readable tables."""
    summary, err_summary = _read_summary()
    details, err_details = _read_details()
    movements_xls, err_mov = _read_movements_xls()
    analysis_xls, err_ana = _read_analysis_xls()

    errors = [e for e in [err_summary, err_details, err_mov, err_ana] if e]

    # Group details by product for easier display
    details_by_product = {}
    for row in details:
        details_by_product.setdefault(row['product'], []).append(row)

    return render(request, 'warehouse/dora_report.html', {
        'summary': summary,
        'details_by_product': details_by_product,
        'movements_xls': movements_xls,
        'analysis_xls': analysis_xls,
        'errors': errors,
        'folder': DATA_FOLDER,
    })


@login_required
def dora_import(request):
    """View: run the import_dora_data management command via web."""
    if request.method == 'POST':
        from django.core.management import call_command
        from io import StringIO
        out = StringIO()
        try:
            call_command('import_dora_data', stdout=out)
            output = out.getvalue()
            messages.success(request, 'Импорт выполнен успешно!')
        except Exception as exc:
            output = str(exc)
            messages.error(request, f'Ошибка при импорте: {exc}')
        return render(request, 'warehouse/dora_import_result.html', {'output': output})

    return render(request, 'warehouse/dora_import.html')
