"""
Management command: import_dora_data
Reads Excel files from "дора лайн движение" folder and imports data into the DB.

Usage:
    python manage.py import_dora_data
    python manage.py import_dora_data --dry-run
"""
import os
import re
from datetime import date, datetime
from decimal import Decimal

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.utils import timezone

import openpyxl
import xlrd

from warehouse.models import (
    Warehouse, ProductCategory, Product, Batch, StockMovement
)

User = get_user_model()

DATA_FOLDER = os.path.join(settings.BASE_DIR, 'дора лайн движение')

FILE_SUMMARY  = 'Доро Лайн движ.xlsx'
FILE_DETAILS  = 'Доро Лайн движ детали.xlsx'
FILE_MOVEMENT = 'Движение номен-ры по фирме ИП Арапова за февраль.xls'
FILE_ANALYSIS = 'Анализ номен-ры по фирме ИП Арапова за февраль.xls'


def _clean(v):
    if v is None:
        return ''
    return str(v).strip()


def _int(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _decimal(v, default=Decimal('0')):
    try:
        return Decimal(str(v))
    except Exception:
        return default


class Command(BaseCommand):
    help = 'Импорт данных из Excel-файлов папки "дора лайн движение"'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Показать что будет импортировано, без записи в БД'
        )

    def handle(self, *args, **options):
        dry = options['dry_run']
        if dry:
            self.stdout.write(self.style.WARNING('=== DRY-RUN (изменения не сохраняются) ==='))

        # Ensure system objects exist
        admin = User.objects.filter(is_superuser=True).first()
        category, _ = ProductCategory.objects.get_or_create(name='Доролайн')
        warehouse, _ = Warehouse.objects.get_or_create(
            name='Основной Склад',
            defaults={'address': 'ИП Арапова'}
        )

        summary = self._parse_summary()
        details = self._parse_details()
        movements_xls = self._parse_movements_xls()
        analysis_xls = self._parse_analysis_xls()

        self.stdout.write(f'\nСводка: {len(summary)} товаров')
        self.stdout.write(f'Детали: {len(details)} транзакций')
        self.stdout.write(f'Движение (ИП Арапова): {len(movements_xls)} строк')
        self.stdout.write(f'Анализ (ИП Арапова): {len(analysis_xls)} товаров')

        if dry:
            self._print_summary(summary)
            self._print_details(details)
            self._print_movements_xls(movements_xls)
            self._print_analysis_xls(analysis_xls)
            return

        # ---- Import products + batches from summary ----
        created_products = 0
        created_batches = 0
        created_movements = 0

        # Build expiry map from analysis file (product_name -> expiry_date)
        expiry_map = {row['product']: row['expiry_date'] for row in analysis_xls if row.get('expiry_date')}

        for row in summary:
            name = row['product']
            sku = re.sub(r'\s+', '_', name.upper())[:100]

            product, p_created = Product.objects.get_or_create(
                sku=sku,
                defaults={
                    'name': name,
                    'category': category,
                    'unit': row.get('unit', 'шт'),
                }
            )
            if p_created:
                created_products += 1
                self.stdout.write(f'  + Товар: {name}')

            # Default expiry: 1 year from now if not in analysis
            exp_date = expiry_map.get(name) or expiry_map.get(name.split()[0])
            if not exp_date:
                exp_date = date(2027, 1, 1)

            # Opening stock batch
            opening = _int(row.get('opening'))
            if opening > 0:
                batch, b_created = Batch.objects.get_or_create(
                    product=product,
                    warehouse=warehouse,
                    batch_number='НАЧАЛЬНЫЙ_ОСТАТОК',
                    defaults={
                        'expiry_date': exp_date,
                        'quantity': opening,
                        'purchase_price': Decimal('0'),
                    }
                )
                if b_created:
                    created_batches += 1
                    StockMovement.objects.create(
                        movement_type='in',
                        batch=batch,
                        warehouse_to=warehouse,
                        quantity=opening,
                        price=Decimal('0'),
                        employee=admin,
                        notes='Начальный остаток (импорт из Excel)',
                        date=timezone.make_aware(datetime(2026, 2, 1)),
                    )
                    created_movements += 1

            # Incoming batch
            incoming = _int(row.get('incoming'))
            if incoming > 0:
                batch_in, bi_created = Batch.objects.get_or_create(
                    product=product,
                    warehouse=warehouse,
                    batch_number='ПРИХОД_ФЕВ_2026',
                    defaults={
                        'expiry_date': exp_date,
                        'quantity': incoming,
                        'purchase_price': Decimal('0'),
                    }
                )
                if bi_created:
                    created_batches += 1
                if not bi_created:
                    # Update quantity
                    batch_in.quantity = incoming
                    batch_in.save()
                StockMovement.objects.get_or_create(
                    movement_type='in',
                    batch=batch_in,
                    warehouse_to=warehouse,
                    quantity=incoming,
                    notes='Приход (импорт из Excel)',
                    defaults={
                        'price': Decimal('0'),
                        'employee': admin,
                        'date': timezone.make_aware(datetime(2026, 2, 14)),
                    }
                )

            # Outgoing
            outgoing = _int(row.get('outgoing'))
            if outgoing > 0:
                # Use incoming batch or opening batch
                batch_out = Batch.objects.filter(
                    product=product, warehouse=warehouse
                ).order_by('-quantity').first()
                if batch_out:
                    StockMovement.objects.get_or_create(
                        movement_type='out',
                        batch=batch_out,
                        warehouse_from=warehouse,
                        quantity=outgoing,
                        notes='Расход (импорт из Excel)',
                        defaults={
                            'price': Decimal('0'),
                            'employee': admin,
                            'date': timezone.make_aware(datetime(2026, 3, 1)),
                        }
                    )
                    created_movements += 1

            # Set closing stock
            closing = _int(row.get('closing'))
            batch_main = Batch.objects.filter(
                product=product, warehouse=warehouse
            ).order_by('-quantity').first()
            if batch_main and closing >= 0:
                batch_main.quantity = closing
                batch_main.save()

        self.stdout.write(self.style.SUCCESS(
            f'\n✅ Импорт завершён: '
            f'{created_products} товаров, {created_batches} партий, {created_movements} движений'
        ))

    # ------------------------------------------------------------------ #
    # Parsers                                                              #
    # ------------------------------------------------------------------ #

    def _parse_summary(self):
        """
        Доро Лайн движ.xlsx
        Columns (row 8): Артикул | Номенклатура | Ед.изм. | Нач.ост. | Приход | Расход | Кон.ост.
        Data starts ~row 10, ends before 'Итого'.
        """
        path = os.path.join(DATA_FOLDER, FILE_SUMMARY)
        if not os.path.exists(path):
            self.stderr.write(f'Файл не найден: {path}')
            return []

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        results = []
        for row in rows:
            name = _clean(row[3]) if len(row) > 3 else ''
            if not name or name in ('Номенклатура', 'Склад', 'Артикул') or name.startswith('Итого'):
                continue
            # Columns: [0]=артикул [3]=номенклатура [8]=ед.изм [9]=нач.ост [13]=приход [15]=расход [16]=кон.ост
            results.append({
                'product': name,
                'unit': _clean(row[8]) if len(row) > 8 else 'шт',
                'opening': row[9] if len(row) > 9 else 0,
                'incoming': row[13] if len(row) > 13 else 0,
                'outgoing': row[15] if len(row) > 15 else 0,
                'closing': row[16] if len(row) > 16 else 0,
            })
        return results

    def _parse_details(self):
        """
        Доро Лайн движ детали.xlsx
        Per-product block: product header row (col 12), then transaction rows (col 0 = doc, col 10 = partner)
        Columns: [12]=product [13]=unit [14]=нач [15]=приход [16]=расход [17]=кон
        Transaction rows: [0]=doc_name [10]=partner [14..17]=qty changes
        """
        path = os.path.join(DATA_FOLDER, FILE_DETAILS)
        if not os.path.exists(path):
            self.stderr.write(f'Файл не найден: {path}')
            return []

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        results = []
        current_product = None
        for row in rows:
            if len(row) < 14:
                continue
            # Product header row: col 12 has product name, col 0 is empty
            product_name = _clean(row[12]) if row[12] else ''
            doc_name = _clean(row[0]) if row[0] else ''

            if product_name and not doc_name:
                current_product = product_name
                continue

            if doc_name and current_product and (
                doc_name.startswith('Расходный') or doc_name.startswith('Приходный')
            ):
                # Parse date from document name
                dt = _parse_doc_date(doc_name)
                partner = _clean(row[10]) if len(row) > 10 else ''
                qty_before = row[14] if len(row) > 14 else None
                qty_in = row[15] if len(row) > 15 else None
                qty_out = row[16] if len(row) > 16 else None
                qty_after = row[17] if len(row) > 17 else None
                results.append({
                    'product': current_product,
                    'doc': doc_name,
                    'date': dt,
                    'partner': partner,
                    'qty_before': _int(qty_before),
                    'qty_in': _int(qty_in),
                    'qty_out': _int(qty_out),
                    'qty_after': _int(qty_after),
                    'type': 'in' if qty_in else 'out',
                })
        return results

    def _parse_movements_xls(self):
        """
        Движение номен-ры по фирме ИП Арапова за февраль.xls
        Columns: [0]=Регистратор [1]=Контрагент [2]=Номенклатура [3]=Приход [4]=Расход
        Data rows start from row 6 (index 5).
        """
        path = os.path.join(DATA_FOLDER, FILE_MOVEMENT)
        if not os.path.exists(path):
            self.stderr.write(f'Файл не найден: {path}')
            return []

        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        results = []
        for i in range(5, ws.nrows):
            row = ws.row_values(i)
            doc = _clean(row[0])
            partner = _clean(row[1])
            product = _clean(row[2])
            if not product:
                continue
            results.append({
                'doc': doc,
                'partner': partner,
                'product': product,
                'incoming': _int(row[3]),
                'outgoing': _int(row[4]),
                'date': _parse_doc_date(doc),
            })
        return results

    def _parse_analysis_xls(self):
        """
        Анализ номен-ры по фирме ИП Арапова за февраль.xls
        Columns: [1]=Номенклатура [3]=Срок [4]=ЕИ [5]=Нач.ост [7]=Приход [9]=Расход [11]=Кон.ост
        Data rows ~7-20 (index 6+).
        """
        path = os.path.join(DATA_FOLDER, FILE_ANALYSIS)
        if not os.path.exists(path):
            self.stderr.write(f'Файл не найден: {path}')
            return []

        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        results = []
        for i in range(6, ws.nrows):
            row = ws.row_values(i)
            product = _clean(row[1]) if len(row) > 1 else ''
            if not product or product.startswith('Итого') or product.startswith('Ед.'):
                continue
            # Parse expiry date
            exp_str = _clean(row[3]) if len(row) > 3 else ''
            exp_date = _parse_date_str(exp_str)
            results.append({
                'product': product,
                'expiry_date': exp_date,
                'unit': _clean(row[4]) if len(row) > 4 else 'шт',
                'opening': _int(row[5]) if len(row) > 5 else 0,
                'incoming': _int(row[7]) if len(row) > 7 else 0,
                'outgoing': _int(row[9]) if len(row) > 9 else 0,
                'closing': _int(row[11]) if len(row) > 11 else 0,
            })
        return results

    # ------------------------------------------------------------------ #
    # Dry-run printers                                                     #
    # ------------------------------------------------------------------ #

    def _print_summary(self, rows):
        self.stdout.write('\n--- Сводка по товарам (Доро Лайн движ.xlsx) ---')
        self.stdout.write(f'{"Товар":<45} {"Ед":>4} {"Нач":>6} {"Прих":>6} {"Расх":>6} {"Кон":>6}')
        for r in rows:
            self.stdout.write(
                f'{r["product"]:<45} {r["unit"]:>4} '
                f'{str(r["opening"] or 0):>6} {str(r["incoming"] or 0):>6} '
                f'{str(r["outgoing"] or 0):>6} {str(r["closing"] or 0):>6}'
            )

    def _print_details(self, rows):
        self.stdout.write(f'\n--- Детальные движения (показано первые 10) ---')
        for r in rows[:10]:
            self.stdout.write(
                f'  {r["date"] or "?"} | {r["product"][:30]:<30} | '
                f'+{r["qty_in"]:>4} -{r["qty_out"]:>4} | {r["partner"][:40]}'
            )

    def _print_movements_xls(self, rows):
        self.stdout.write(f'\n--- Движение ИП Арапова (первые 10) ---')
        for r in rows[:10]:
            self.stdout.write(
                f'  {r["date"] or "?"} | {r["product"][:30]:<30} | '
                f'+{r["incoming"]:>4} -{r["outgoing"]:>4} | {r["partner"][:35]}'
            )

    def _print_analysis_xls(self, rows):
        self.stdout.write('\n--- Анализ номенклатуры ИП Арапова ---')
        self.stdout.write(f'{"Товар":<45} {"Срок":>12} {"Нач":>6} {"Прих":>6} {"Расх":>6} {"Кон":>6}')
        for r in rows:
            self.stdout.write(
                f'{r["product"]:<45} {str(r["expiry_date"] or "—"):>12} '
                f'{r["opening"]:>6} {r["incoming"]:>6} {r["outgoing"]:>6} {r["closing"]:>6}'
            )


# ------------------------------------------------------------------ #
# Helpers                                                              #
# ------------------------------------------------------------------ #

def _parse_doc_date(doc_str):
    """Extract date from a document string like 'Расходный ордер ... от 05.02.2026 10:06:00'"""
    m = re.search(r'от\s+(\d{2}\.\d{2}\.\d{4})', doc_str)
    if m:
        try:
            return datetime.strptime(m.group(1), '%d.%m.%Y').date()
        except ValueError:
            pass
    return None


def _parse_date_str(s):
    """Parse a date string like '01.02.2027'"""
    s = s.strip()
    if not s:
        return None
    for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None
